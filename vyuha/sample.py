"""Generate a deliberately messy demo workbook.

Two jobs. It is the fixture the tests run against, and it is the file the
founder opens a demo with when a prospect has not yet handed over their own
data. Every ugly thing in here is copied from a real distributor export:

* three junk rows and a merged title above the header
* rupee symbols, thousands commas, "(4,500)" negatives and "5,000 Cr" credits
* blank spacer rows in the middle of the data
* a "Grand Total" row that must not be counted
* the same customer spelled four different ways
* dates as real dates in one sheet and as dd-mm-yyyy text in another
* three SKUs sitting in stock that never appear in sales (dead stock)
* one fast-moving SKU with barely a week of cover left
"""

from __future__ import annotations

import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

FIRM = "Shree Balaji Distributors"

# (sku, name, category, rate)
PRODUCTS: tuple[tuple[str, str, str, float], ...] = (
    ("BRG-6205", "Ball Bearing 6205 ZZ", "Bearings", 148.0),
    ("BRG-6301", "Ball Bearing 6301 2RS", "Bearings", 172.0),
    ("BLT-A45", "V-Belt A45 Industrial", "Belts", 310.0),
    ("BLT-B60", "V-Belt B60 Heavy Duty", "Belts", 465.0),
    ("SEA-25X40", "Oil Seal 25x40x7", "Seals", 62.0),
    ("SEA-35X52", "Oil Seal 35x52x10", "Seals", 88.0),
    ("PMP-CI15", "CI Pump Casing 1.5in", "Pumps", 2450.0),
    ("PMP-IMP20", "Pump Impeller 2in Bronze", "Pumps", 3180.0),
    ("VLV-BR15", "Brass Ball Valve 15mm", "Valves", 415.0),
    ("VLV-BR25", "Brass Ball Valve 25mm", "Valves", 720.0),
    ("GRS-EP2", "EP2 Lithium Grease 1kg", "Lubricants", 340.0),
    ("OIL-HYD68", "Hydraulic Oil 68 - 20L", "Lubricants", 3650.0),
    ("CHN-08B", "Roller Chain 08B x 5ft", "Chains", 890.0),
    ("CHN-12B", "Roller Chain 12B x 5ft", "Chains", 1480.0),
    ("CPL-L095", "Flexible Coupling L095", "Couplings", 1250.0),
)

# Customers, with the messy spellings a ledger accumulates over the years.
CUSTOMERS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("Sharma Traders", ("M/s Sharma Traders", "Sharma traders", "SHARMA TRADERS",
                        "M/s. Sharma Traders Pvt Ltd")),
    ("Gupta Engineering Works", ("Gupta Engineering Works", "gupta engineering works ",
                                 "M/s Gupta Engineering Works")),
    ("Krishna Auto Spares", ("Krishna Auto Spares", "KRISHNA AUTO SPARES")),
    ("Vikram Industrial Supply", ("Vikram Industrial Supply", "Vikram Industrial Supply Co.")),
    ("Deepak Machinery", ("Deepak Machinery", "M/s Deepak Machinery")),
    ("Anand Hardware", ("Anand Hardware",)),
    ("Patel Pumps & Motors", ("Patel Pumps & Motors", "patel pumps and motors")),
    ("Modern Tools Centre", ("Modern Tools Centre",)),
    ("Jain Bearing House", ("Jain Bearing House", "JAIN BEARING HOUSE")),
    ("Reliable Spares", ("Reliable Spares",)),
)

# SKUs that will never appear in the sales sheet -> dead stock.
DEAD_SKUS = frozenset({"PMP-IMP20", "CHN-12B", "VLV-BR25"})
# SKU deliberately given a tiny stock balance against heavy sales -> low cover.
FAST_MOVER = "BRG-6205"

WAREHOUSES = ("Main Godown", "Shop Counter", "Godown 2")


def build(path: str | Path, as_of: datetime | None = None, seed: int = 7) -> Path:
    """Write the demo workbook to ``path`` and return the path."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    as_of = as_of or datetime.now()
    rng = random.Random(seed)

    book = Workbook()
    sales_rows = _write_sales(book, rng, as_of)
    _write_stock(book, rng, sales_rows)
    _write_outstanding(book, rng, as_of, sales_rows)
    _write_junk_sheet(book)

    book.save(path)
    return path


# --- sheets ---------------------------------------------------------------


def _write_sales(book: Workbook, rng: random.Random, as_of: datetime) -> list[dict]:
    ws = book.active
    ws.title = "Sales Register"

    # Junk header block, complete with a merged title cell.
    ws["A1"] = FIRM
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = "Plot 14, MIDC Industrial Area, Pune 411026   GSTIN: 27AABCS1429P1ZQ"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws["A3"] = "SALES REGISTER"
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)
    # Row 4 left blank on purpose.

    headers = ["Bill Date", "Invoice No.", "Party Name", "Item Code",
               "Item Description", "Qty (Nos)", "Rate", "Amount"]
    for column, name in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=column, value=name)
        cell.font = Font(bold=True)

    sellable = [p for p in PRODUCTS if p[0] not in DEAD_SKUS]
    rows: list[dict] = []
    row_number = 6
    invoice_number = 2041
    start = as_of - timedelta(days=180)

    for day_offset in range(0, 181):
        date = start + timedelta(days=day_offset)
        if date.weekday() == 6:  # closed Sundays
            continue
        if rng.random() < 0.35:
            continue
        for _ in range(rng.randint(1, 3)):
            invoice_number += 1
            canonical, spellings = CUSTOMERS[rng.randrange(len(CUSTOMERS))]
            party = spellings[rng.randrange(len(spellings))]
            invoice_id = f"INV-{invoice_number}"
            for _ in range(rng.randint(1, 3)):
                sku, name, _category, rate = (
                    next(p for p in PRODUCTS if p[0] == FAST_MOVER)
                    if rng.random() < 0.18
                    else sellable[rng.randrange(len(sellable))]
                )
                qty = rng.randint(1, 14)
                # The rate a clerk types drifts from the list price.
                actual_rate = round(rate * rng.uniform(0.94, 1.09), 2)
                amount = round(qty * actual_rate, 2)

                ws.cell(row=row_number, column=1, value=date)
                ws.cell(row=row_number, column=1).number_format = "DD-MM-YYYY"
                ws.cell(row=row_number, column=2, value=invoice_id)
                ws.cell(row=row_number, column=3, value=party)
                ws.cell(row=row_number, column=4, value=sku)
                ws.cell(row=row_number, column=5, value=name)
                ws.cell(row=row_number, column=6, value=qty)
                # Some cells come through as formatted text, not numbers.
                if rng.random() < 0.12:
                    ws.cell(row=row_number, column=7, value=f"₹ {actual_rate:,.2f}")
                    ws.cell(row=row_number, column=8, value=f"{amount:,.2f}")
                else:
                    ws.cell(row=row_number, column=7, value=actual_rate)
                    ws.cell(row=row_number, column=8, value=amount)

                rows.append({
                    "date": date, "invoice": invoice_id, "party": party,
                    "canonical": canonical, "sku": sku, "qty": qty, "amount": amount,
                })
                row_number += 1

            # Spacer rows between invoices, exactly as a human would leave them.
            if rng.random() < 0.06:
                row_number += 1

    total = sum(r["amount"] for r in rows)
    row_number += 1
    ws.cell(row=row_number, column=5, value="Grand Total").font = Font(bold=True)
    ws.cell(row=row_number, column=6, value=sum(r["qty"] for r in rows))
    ws.cell(row=row_number, column=8, value=round(total, 2)).font = Font(bold=True)

    _widen(ws, headers)
    return rows


def _write_stock(book: Workbook, rng: random.Random, sales_rows: list[dict]) -> None:
    ws = book.create_sheet("Stock Statement")

    ws["A1"] = f"{FIRM} — Stock as on date"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws["A1"].font = Font(bold=True, size=12)

    headers = ["Item Code", "Product Name", "Category", "Closing Stock",
               "Reorder Level", "Purchase Rate", "Godown"]
    for column, name in enumerate(headers, start=1):
        ws.cell(row=3, column=column, value=name).font = Font(bold=True)

    sold_qty: dict[str, int] = {}
    for row in sales_rows:
        sold_qty[row["sku"]] = sold_qty.get(row["sku"], 0) + row["qty"]

    row_number = 4
    for sku, name, category, rate in PRODUCTS:
        moved = sold_qty.get(sku, 0)
        if sku == FAST_MOVER:
            stock = max(int(moved / 180 * 6), 3)  # about six days of cover
            reorder = max(int(moved / 180 * 30), 25)
        elif sku in DEAD_SKUS:
            stock = rng.randint(18, 55)
            reorder = rng.randint(4, 10)
        else:
            stock = rng.randint(0, 140)
            reorder = rng.randint(10, 40)

        ws.cell(row=row_number, column=1, value=sku)
        ws.cell(row=row_number, column=2, value=name)
        ws.cell(row=row_number, column=3, value=category)
        ws.cell(row=row_number, column=4, value=stock)
        ws.cell(row=row_number, column=5, value=reorder)
        ws.cell(row=row_number, column=6, value=rate)
        ws.cell(row=row_number, column=7, value=WAREHOUSES[rng.randrange(len(WAREHOUSES))])
        row_number += 1

    # A duplicated row, because someone pasted twice.
    for column in range(1, 8):
        ws.cell(row=row_number, column=column,
                value=ws.cell(row=row_number - 1, column=column).value)

    _widen(ws, headers)


def _write_outstanding(
    book: Workbook, rng: random.Random, as_of: datetime, sales_rows: list[dict]
) -> None:
    ws = book.create_sheet("Outstanding")

    headers = ["Party", "Bill No", "Bill Date", "Due Date", "Balance Due", "Remarks"]
    for column, name in enumerate(headers, start=1):
        ws.cell(row=1, column=column, value=name).font = Font(bold=True)

    # Leave the most recent invoices unpaid, plus a long tail of old ones.
    by_date = sorted(sales_rows, key=lambda r: r["date"], reverse=True)
    seen_invoices: set[str] = set()
    candidates: list[dict] = []
    for row in by_date:
        if row["invoice"] in seen_invoices:
            continue
        seen_invoices.add(row["invoice"])
        candidates.append(row)

    picked = candidates[:26] + [c for c in candidates[26:] if rng.random() < 0.06][:14]

    row_number = 2
    for entry in picked:
        age = (as_of.date() - entry["date"].date()).days
        credit_days = 30 if age < 120 else 45
        due = entry["date"] + timedelta(days=credit_days)
        balance = round(entry["amount"] * rng.uniform(0.4, 1.0), 2)

        ws.cell(row=row_number, column=1, value=entry["party"])
        ws.cell(row=row_number, column=2, value=entry["invoice"])
        # Dates as dd-mm-yyyy *text* here — a different export, a different habit.
        ws.cell(row=row_number, column=3, value=entry["date"].strftime("%d-%m-%Y"))
        ws.cell(row=row_number, column=4, value=due.strftime("%d-%m-%Y"))

        roll = rng.random()
        if roll < 0.10:
            ws.cell(row=row_number, column=5, value=f"₹{balance:,.2f}")
        elif roll < 0.14:
            ws.cell(row=row_number, column=5, value=f"({balance:,.2f})")
            ws.cell(row=row_number, column=6, value="credit note")
        elif roll < 0.18:
            ws.cell(row=row_number, column=5, value=f"{balance:,.2f} Cr")
            ws.cell(row=row_number, column=6, value="advance received")
        else:
            ws.cell(row=row_number, column=5, value=balance)
        if age > 150:
            ws.cell(row=row_number, column=6, value="follow up - legal notice sent")
        row_number += 1

    ws.cell(row=row_number + 1, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=row_number + 1, column=5, value="=SUM(E2:E{})".format(row_number - 1))

    _widen(ws, headers)


def _write_junk_sheet(book: Workbook) -> None:
    """A sheet of notes — the pipeline should recognise it as unusable and skip it."""
    ws = book.create_sheet("Notes")
    ws["A1"] = "Reminders"
    ws["A2"] = "Call transporter about Nashik delivery"
    ws["A3"] = "GST filing 20th"
    ws["A5"] = "New rate list from principal expected next month"


def _widen(ws, headers: list[str]) -> None:
    for index, name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(index)].width = max(len(name) + 4, 14)
