"""Read whatever the client sends us into raw cell grids.

Deliberately dumb: this layer does not decide what anything *means*. It opens
.xlsx/.xlsm/.xls/.csv, unmerges merged cells by spreading the anchor value
across the range (the single most common wrecker of Excel exports), and hands
back one raw grid per sheet with no header assumed. Interpretation happens in
:mod:`vyuha.detect`.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd

EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
LEGACY_EXCEL_SUFFIXES = {".xls"}
CSV_SUFFIXES = {".csv", ".txt", ".tsv"}


class IngestError(RuntimeError):
    """The file could not be opened or contained nothing usable."""


@dataclass
class RawSheet:
    """One sheet, as a grid of raw cell values with no header applied."""

    name: str
    grid: pd.DataFrame
    source: Path
    # grid row index -> 1-based row number in the original file, so we can tell
    # the client "your header is on row 5" after blank rows have been trimmed.
    row_map: list[int] = field(default_factory=list)

    @property
    def is_empty(self) -> bool:
        return self.grid.empty

    def source_row(self, grid_row: int) -> int:
        if 0 <= grid_row < len(self.row_map):
            return self.row_map[grid_row]
        return grid_row + 1


def read_source(path: str | Path) -> list[RawSheet]:
    """Open ``path`` and return one :class:`RawSheet` per non-empty sheet."""
    path = Path(path)
    if not path.exists():
        raise IngestError(f"File not found: {path}")
    if path.is_dir():
        raise IngestError(f"Expected a file, got a directory: {path}")

    suffix = path.suffix.lower()
    if suffix in CSV_SUFFIXES:
        sheets = [_read_csv(path)]
    elif suffix in EXCEL_SUFFIXES:
        sheets = _read_xlsx(path)
    elif suffix in LEGACY_EXCEL_SUFFIXES:
        sheets = _read_legacy_excel(path)
    else:
        raise IngestError(
            f"Unsupported file type '{suffix}'. Send an .xlsx, .xls or .csv file."
        )

    sheets = [s for s in sheets if not s.is_empty]
    if not sheets:
        raise IngestError(f"No data found in {path.name} — every sheet was blank.")
    return sheets


# --- excel ----------------------------------------------------------------


def _read_xlsx(path: Path) -> list[RawSheet]:
    from openpyxl import load_workbook

    try:
        # data_only=True gives us the cached result of formulas rather than
        # "=SUM(B2:B9)". Files saved by a tool that never calculated will have
        # None there; nothing we can do about that without a spreadsheet engine.
        book = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:  # openpyxl raises a zoo of exception types
        raise IngestError(f"Could not open {path.name}: {exc}") from exc

    sheets: list[RawSheet] = []
    try:
        for ws in book.worksheets:
            if ws.sheet_state != "visible":
                continue  # hidden sheets are almost always scratch work
            rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            if not rows:
                continue
            grid = pd.DataFrame(rows)
            grid = _spread_merged_cells(grid, ws.merged_cells.ranges)
            trimmed, row_map = _trim(grid)
            sheets.append(
                RawSheet(name=str(ws.title), grid=trimmed, source=path, row_map=row_map)
            )
    finally:
        book.close()
    return sheets


def _read_legacy_excel(path: Path) -> list[RawSheet]:
    try:
        book = pd.read_excel(path, sheet_name=None, header=None, dtype=object)
    except ImportError as exc:
        raise IngestError(
            f"{path.name} is a legacy .xls file. Install `xlrd` "
            "(pip install xlrd) or re-save it as .xlsx."
        ) from exc
    except Exception as exc:
        raise IngestError(f"Could not open {path.name}: {exc}") from exc
    sheets: list[RawSheet] = []
    for name, frame in book.items():
        trimmed, row_map = _trim(frame.reset_index(drop=True))
        sheets.append(RawSheet(name=str(name), grid=trimmed, source=path, row_map=row_map))
    return sheets


def _spread_merged_cells(grid: pd.DataFrame, ranges) -> pd.DataFrame:
    """Copy each merged range's anchor value into every cell of the range.

    Excel stores a merged block as one value plus a pile of ``None``s. Left
    alone, that turns a category header spanning six rows into five blank
    rows. Coordinates from openpyxl are 1-based; the grid is 0-based.
    """
    for cell_range in list(ranges):
        r0, c0 = cell_range.min_row - 1, cell_range.min_col - 1
        r1, c1 = cell_range.max_row - 1, cell_range.max_col - 1
        if r0 >= len(grid.index) or c0 >= len(grid.columns):
            continue
        value = grid.iat[r0, c0]
        if value is None:
            continue
        r1 = min(r1, len(grid.index) - 1)
        c1 = min(c1, len(grid.columns) - 1)
        grid.iloc[r0 : r1 + 1, c0 : c1 + 1] = value
    return grid


# --- csv ------------------------------------------------------------------


def _read_csv(path: Path) -> RawSheet:
    delimiter = _sniff_delimiter(path)
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            grid = pd.read_csv(
                path,
                header=None,
                dtype=object,
                sep=delimiter,
                encoding=encoding,
                skip_blank_lines=False,
                engine="python",
            )
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
        except Exception as exc:
            raise IngestError(f"Could not read {path.name}: {exc}") from exc
        trimmed, row_map = _trim(grid)
        return RawSheet(name=path.stem, grid=trimmed, source=path, row_map=row_map)
    raise IngestError(f"Could not decode {path.name}: {last_error}")


def _sniff_delimiter(path: Path) -> str:
    try:
        with path.open("r", encoding="utf-8-sig", errors="replace") as fh:
            sample = fh.read(8192)
        if not sample.strip():
            return ","
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except (csv.Error, OSError):
        return "\t" if path.suffix.lower() == ".tsv" else ","


# --- shared ---------------------------------------------------------------


def _trim(grid: pd.DataFrame) -> tuple[pd.DataFrame, list[int]]:
    """Drop all-blank rows and columns; return the grid and its row numbering.

    Interior blank rows are dropped too — in an Excel export they are visual
    spacing, never data — but the row *order* is preserved so a totals row at
    the bottom still lands at the bottom. The returned list maps each surviving
    grid row back to its 1-based row number in the original file.
    """
    if grid.empty:
        return grid, []
    blank = grid.map(_is_blank)
    grid = grid.loc[~blank.all(axis=1), ~blank.all(axis=0)]
    row_map = [int(label) + 1 for label in grid.index]
    grid = grid.reset_index(drop=True).T.reset_index(drop=True).T
    return grid, row_map


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and value != value:  # NaN
        return True
    return isinstance(value, str) and not value.strip()
