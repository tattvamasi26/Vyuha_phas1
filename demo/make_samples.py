"""Generate the demo corpus — the messy files a real client actually sends.

    .venv/Scripts/python demo/make_samples.py

``vyuha/sample.py`` already builds one messy workbook, and it is good, but it
demonstrates *one* kind of mess. A prospect's objection is never "does it read a
spreadsheet" — it is "you have not seen **my** files". So this builds nine, each
carrying a different failure the engine has to survive, and each labelled with
what it proves.

The files are **generated, never hand-made**, for the same reason the demo
tenant is seeded rather than typed: binaries in a repo go stale, nobody
remembers what is inside them, and a demo that depends on a file somebody edited
in 2026 is a demo that breaks. Re-run this and the corpus is exactly what the
README says it is.

Dates are relative to today, so the "recent" file is always recent and a demo
never opens on data from last year.

Every file here is mess the engine is *expected to survive*. The one exception
is ``09-broken.xlsx``, which must fail — cleanly, with a message a person can
act on. A product that only ever demonstrates success teaches nobody what
happens on a bad day, and the honest failure is worth showing.
"""

from __future__ import annotations

import csv
import random
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = HERE / "samples"
SEED = 20260831

ITEMS = [
    ("URE-50", "Urea 50kg", "Fertiliser", "bag", 320, 268),
    ("DAP-50", "DAP 50kg", "Fertiliser", "bag", 1420, 1250),
    ("POT-50", "Potash 50kg", "Fertiliser", "bag", 980, 860),
    ("GYP-05", "Gypsum 5kg", "Fertiliser", "bag", 140, 98),
    ("NEM-10", "Neem Cake 10kg", "Fertiliser", "bag", 380, 305),
    ("CTF-50", "Cattle Feed 50kg", "Feed", "bag", 1180, 1040),
    ("PDY-05", "Paddy Seed 5kg", "Seed", "pkt", 640, 520),
    ("SPR-16", "Sprayer 16L", "Tools", "piece", 2250, 1780),
    ("HDP-10", "HDPE Pipe 1in 10m", "Hardware", "coil", 870, 690),
    ("TRP-12", "Tarpaulin 12x15", "Hardware", "piece", 1150, 900),
]

PARTIES = ["Ramu Stores", "M/s Ramu Stores", "Ramu Stores.", "RAMU STORES",
           "Basavaraj Agri Centre", "Krishna Traders Pvt Ltd", "Shetty Farms",
           "Patil Nursery", "Hanumanth & Sons", "Cash"]


def ago(days: int) -> date:
    return date.today() - timedelta(days=days)


def _autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        width = max((len(str(ws.cell(r, col).value or ""))
                     for r in range(1, min(ws.max_row, 40) + 1)), default=8)
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 4, 11), 30)


# ------------------------------------------------------------------ 01 clean

def clean_csv(rng: random.Random) -> tuple[str, str]:
    """The easy case. Proves nothing on its own — it is the control."""
    path = OUT / "01-clean-sales.csv"
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["Date", "Invoice No", "Party Name", "Item Code", "Item Name",
                    "Quantity", "Rate", "Amount"])
        for i in range(60):
            sku, name, _, _, rate, _ = rng.choice(ITEMS)
            qty = rng.choice([5, 10, 20, 25, 40])
            w.writerow([ago(rng.randint(1, 120)).strftime("%Y-%m-%d"),
                        f"INV-{2400 + i}", rng.choice(PARTIES[:8]), sku, name,
                        qty, rate, qty * rate])
    return path.name, "A clean CSV. The control — if this fails, nothing else matters."


# ------------------------------------------------- 02 the filthy workbook

def filthy_workbook(rng: random.Random) -> tuple[str, str]:
    """Everything wrong at once, in one file, across three sheets."""
    path = OUT / "02-filthy-multisheet.xlsx"
    wb = Workbook()

    # -- sheet 1: junk rows, a merged title, blank spacers, a grand total
    ws = wb.active
    ws.title = "Sales Register"
    ws["A1"] = "SHREE AGRO & HARDWARE — BELAGAVI"
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = "GSTIN: 29ABCDE1234F1Z5   |   Ph: 0831-2345678"
    ws.merge_cells("A2:H2")
    ws["A3"] = "Sales Register for the period 01-04-2026 to date"
    ws.merge_cells("A3:H3")
    # row 4 blank, header lands on row 5 — the thing find_header_row must get right
    ws.append([])
    ws.append(["Date", "Bill No", "Customer", "Item", "Qty", "Rate", "Amount", "Remarks"])

    total = 0.0
    for i in range(70):
        sku, name, _, _, rate, _ = rng.choice(ITEMS)
        qty = rng.choice([5, 10, 20, 30, 50])
        amount = qty * rate
        total += amount
        # dates as dd-mm-yyyy TEXT, not real dates
        when = ago(rng.randint(1, 200)).strftime("%d-%m-%Y")
        # amounts in three different dialects
        style = i % 4
        if style == 0:
            cell = f"₹ {amount:,.2f}"
        elif style == 1:
            cell = f"{amount:,}"
        elif style == 2 and amount > 20000:
            cell = f"({amount:,})"          # parentheses = a credit note
            amount = -amount
            total -= 2 * (qty * rate)
        else:
            cell = amount
        ws.append([when, f"B-{1200 + i}", rng.choice(PARTIES), name, qty, rate, cell,
                   rng.choice(["", "", "", "urgent", "part load", "check rate"])])
        if i in (23, 47):
            ws.append([])                    # blank spacer mid-table

    ws.append([])
    ws.append(["", "", "", "Grand Total", "", "", total, ""])   # must be excluded

    # -- sheet 2: stock, with real dates and a location column
    st = wb.create_sheet("Stock Statement")
    st.append(["Closing Stock as on " + date.today().strftime("%d-%m-%Y")])
    st.merge_cells("A1:F1")
    st.append([])
    st.append(["Item Code", "Description", "Godown", "Closing Qty",
               "Reorder Level", "Value"])
    for sku, name, _, _, rate, cost in ITEMS:
        qty = rng.choice([0, 4, 9, 18, 30, 55, 90])
        st.append([sku, name, rng.choice(["Belagavi", "Hubballi"]), qty,
                   rng.choice([10, 15, 20, 25]), qty * cost])

    # -- sheet 3: prose that must be skipped entirely
    notes = wb.create_sheet("Notes")
    notes["A1"] = "Points for the meeting"
    for i, line in enumerate([
        "Coromandel rate revision expected next month.",
        "Hubballi godown shutter repaired on 12th.",
        "Ramu Stores asking for 45 days credit — decide.",
        "Do not reorder soil test kits, no movement since March.",
    ], start=3):
        notes[f"A{i}"] = line

    for sheet in wb.worksheets:
        _autosize(sheet)
    wb.save(path)
    return path.name, ("Junk rows above the header, a merged title, a blank spacer "
                       "mid-table, a Grand Total row, dates as text, ₹ and comma and "
                       "parenthesised amounts, four spellings of one customer, and a "
                       "prose sheet that must be skipped.")


# -------------------------------------------- 03-05 monthly, drifting columns

def monthly_drift(rng: random.Random) -> list[tuple[str, str]]:
    """Three months from the same client, whose columns move between files.

    This is the case nobody designs for and every client produces: the same
    report, exported by a different person each month, with the columns renamed
    and reordered. All three must land as the same kind of table.
    """
    made = []
    shapes = [
        (["Date", "Party", "Item", "Qty", "Rate", "Amount"],
         lambda w, s, n, q, r: [w, s, n, q, r, q * r]),
        # month two: someone added a column and renamed two
        (["Invoice Date", "Customer Name", "Product", "Nos", "Unit Price",
          "Net Amount", "Salesman"],
         lambda w, s, n, q, r: [w, s, n, q, r, q * r,
                                rng.choice(["Mahesh", "Iranna"])]),
        # month three: reordered, and the amount column is gone entirely,
        # so the engine has to derive it from qty x rate
        (["Bill No", "Dt", "Item Description", "Buyer", "Rate", "Quantity"],
         None),
    ]
    for idx, month_back in enumerate((3, 2, 1)):
        headers, build = shapes[idx]
        path = OUT / f"0{3 + idx}-month-{ago(month_back * 30).strftime('%b-%Y').lower()}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(headers)
        for i in range(35):
            sku, name, _, _, rate, _ = rng.choice(ITEMS)
            qty = rng.choice([5, 10, 20, 30])
            when = ago(month_back * 30 + rng.randint(0, 27))
            party = rng.choice(PARTIES[:8])
            if build is not None:
                ws.append(build(when, party, name, qty, rate))
            else:
                ws.append([f"S-{900 + i}", when, name, party, rate, qty])
        _autosize(ws)
        wb.save(path)
        made.append((path.name,
                     ["The baseline monthly export.",
                      "Same report, two columns renamed and one added.",
                      "Columns reordered and the Amount column missing entirely — "
                      "it has to be derived from quantity × rate."][idx]))
    return made


# ------------------------------------------------------- 06 tab-separated

def tsv_export(rng: random.Random) -> tuple[str, str]:
    """What a legacy accounting package spits out. Not a spreadsheet at all."""
    path = OUT / "06-tally-export.txt"
    with path.open("w", encoding="utf-8") as fh:
        fh.write("Particulars\tVch No.\tVch Type\tDebit\tCredit\tOpening\tClosing\n")
        for i in range(40):
            party = rng.choice(PARTIES[:8])
            amount = rng.choice([12800, 45600, 8900, 132000, 24500])
            fh.write(f"{party}\tSL/{2000 + i}\tSales\t{amount}\t\t0\t{amount}\n")
    return path.name, ("A tab-separated export from an accounting package — no "
                       "spreadsheet involved. Delimiter has to be sniffed, and the "
                       "accounting vocabulary (Vch No., Debit) mapped.")


# -------------------------------------------------- 07 price list, 2-row header

def price_list(rng: random.Random) -> tuple[str, str]:
    """A two-row header with merged group cells above it."""
    path = OUT / "07-price-list-merged-header.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Rate List"
    ws["A1"] = "RATE LIST — EFFECTIVE " + date.today().strftime("%d-%m-%Y")
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(bold=True)
    # row 2 is a merged group header sitting ON TOP of the real header in row 3
    ws["C2"] = "Selling"
    ws.merge_cells("C2:D2")
    ws["E2"] = "Stock"
    ws.merge_cells("E2:F2")
    ws["C2"].alignment = Alignment(horizontal="center")
    ws["E2"].alignment = Alignment(horizontal="center")
    ws.append([])
    ws["A3"], ws["B3"] = "Code", "Particulars"
    ws["C3"], ws["D3"] = "MRP", "Dealer Rate"
    ws["E3"], ws["F3"] = "In Godown", "Min Level"
    ws["G3"] = "Unit"
    for sku, name, _, unit, rate, cost in ITEMS:
        ws.append([sku, name, rate, cost, rng.choice([0, 8, 22, 60]),
                   rng.choice([10, 20]), unit])
    _autosize(ws)
    wb.save(path)
    return path.name, ("A price list whose real header is row 3, under a merged "
                       "group header in row 2. The merged cells must not be read "
                       "as the column names.")


# ----------------------------------------------------- 08 WhatsApp transcript

def whatsapp_chat(rng: random.Random) -> tuple[str, str]:
    """An exported WhatsApp thread with real orders buried in conversation.

    Not a table in any sense — this is the file that proves intake is not just
    a spreadsheet reader. Orders, payments and stock notes are mixed in with
    greetings and chatter, in the register people actually type in.
    """
    path = OUT / "08-whatsapp-orders.txt"
    lines = [
        "31/08/2026, 8:12 am - Messages are end-to-end encrypted.",
        "28/08/2026, 9:03 am - Ramu Stores: Namaskara sir",
        "28/08/2026, 9:03 am - Ramu Stores: 20 bags urea beku, rate enu?",
        "28/08/2026, 9:11 am - You: Namaskara. Urea 320 per bag sir",
        "28/08/2026, 9:14 am - Ramu Stores: ok send 20 bags urea and 5 bag gypsum",
        "28/08/2026, 9:15 am - Ramu Stores: tomorrow morning delivery madi",
        "28/08/2026, 9:22 am - You: Done sir, 20 urea + 5 gypsum. Total 7100",
        "28/08/2026, 2:40 pm - Basavaraj Agri Centre: sir 10 bag DAP urgent",
        "28/08/2026, 2:41 pm - Basavaraj Agri Centre: and 2 sprayer 16L",
        "28/08/2026, 3:02 pm - You: DAP 1420, sprayer 2250. Sending today",
        "29/08/2026, 10:15 am - Shetty Farms: Payment done 45000 NEFT",
        "29/08/2026, 10:16 am - Shetty Farms: check and confirm",
        "29/08/2026, 11:40 am - You: Received sir, thank you",
        "29/08/2026, 4:20 pm - Patil Nursery: 30 packet paddy seed and 12 neem cake",
        "30/08/2026, 8:05 am - Krishna Traders Pvt Ltd: potash stock ideya?",
        "30/08/2026, 8:31 am - You: Yes sir 30 bags available",
        "30/08/2026, 8:33 am - Krishna Traders Pvt Ltd: hold 25 bags for me",
        "30/08/2026, 6:12 pm - Hanumanth & Sons: 40 bags cattle feed next week",
        "31/08/2026, 7:50 am - Ramu Stores: sir last month bill pending ide, 12800",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")
    return path.name, ("An exported WhatsApp thread. Five real orders, one payment "
                       "and one outstanding-balance mention, buried in Kannada-English "
                       "chatter. Proves intake is not only a spreadsheet reader.")


# ---------------------------------------------------------- 09 must fail well

def broken_file() -> tuple[str, str]:
    """A file that cannot be read. The message matters more than the failure."""
    path = OUT / "09-broken.xlsx"
    path.write_bytes(b"%PDF-1.4\n% not actually a workbook\n" + b"\x00" * 200)
    return path.name, ("A PDF renamed .xlsx — the single most common real-world "
                       "upload accident. Must fail cleanly and say what to do, "
                       "never crash and never silently produce zeros.")


# ------------------------------------------------------------------ the corpus

def build() -> list[tuple[str, str]]:
    OUT.mkdir(parents=True, exist_ok=True)
    for old in OUT.glob("*"):
        if old.is_file():
            old.unlink()

    rng = random.Random(SEED)
    made: list[tuple[str, str]] = [clean_csv(rng), filthy_workbook(rng)]
    made += monthly_drift(rng)
    made += [tsv_export(rng), price_list(rng), whatsapp_chat(rng), broken_file()]

    readme = ["# Demo corpus",
              "",
              "Generated by `demo/make_samples.py` — **do not hand-edit**, re-run it.",
              "Dates are relative to the day it was generated, so the corpus is never stale.",
              "",
              "Each file carries a different real-world mess:",
              ""]
    for name, why in made:
        readme.append(f"- **`{name}`** — {why}")
    readme += ["",
               "## Using them in a demo",
               "",
               "Upload them one at a time on the client's Add data tab and read the",
               "\"What Vyuha read from your file\" panel out loud. That panel is the pitch:",
               "it names the sheet, the row the header landed on, the columns understood,",
               "and every fix applied.",
               "",
               "Finish on `09-broken.xlsx`. Showing the failure is more convincing than",
               "hiding it — a prospect who has only seen successes does not believe any",
               "of them."]
    (OUT / "README.md").write_text("\n".join(readme) + "\n", encoding="utf-8")
    return made


def main() -> int:
    if "--bulk" in sys.argv:
        i = sys.argv.index("--bulk")
        count = int(sys.argv[i + 1]) if len(sys.argv) > i + 1 else 100
        rows = bulk(count)
        target = HERE / "samples" / "bulk"
        total = sum((target / n).stat().st_size for n, _w in rows)
        print(f"\n  Bulk corpus -> {target}\n  {'-' * 58}")
        kinds: dict[str, int] = {}
        for name, _why in rows:
            kinds[name.rsplit(".", 1)[-1]] = kinds.get(name.rsplit(".", 1)[-1], 0) + 1
        for ext, n in sorted(kinds.items()):
            print(f"  {n:>4} .{ext}")
        print(f"\n  {len(rows)} files, {total / 1_000_000:.1f} MB\n")
        return 0

    made = build()
    print(f"\n  Demo corpus -> {OUT}\n  {'-' * 58}")
    for name, why in made:
        size = (OUT / name).stat().st_size
        print(f"  {name:<38} {size / 1024:6.1f} KB")
    print(f"\n  {len(made)} files, plus README.md\n")
    return 0


# ============================================================ bulk generation

#: Ways a client's own export drifts between months. Real drift, collected from
#: the shapes a distributor's files actually take: somebody renames a column,
#: somebody reorders them, somebody's template drops the total and leaves you to
#: multiply it out.
_HEADER_SHAPES = [
    ["Date", "Party", "Item", "Qty", "Rate", "Amount"],
    ["Invoice Date", "Customer Name", "Product", "Nos", "Unit Price", "Net Amount"],
    ["Bill Dt", "Buyer", "Item Description", "Quantity", "Rate", "Value"],
    ["Dt", "Party Name", "Particulars", "Qty", "Price", "Total"],
    ["Date", "Customer", "Item", "Qty", "Rate"],                 # no amount column
    ["Sl", "Date", "Party", "Item Code", "Item", "Qty", "Rate", "Amount", "Remarks"],
]

_STOCK_SHAPES = [
    ["Item Code", "Description", "Closing Qty", "Reorder Level", "Rate"],
    ["Code", "Particulars", "In Godown", "Min Level", "MRP"],
    ["SKU", "Item Name", "Balance Qty", "ROL", "Selling Price", "Godown"],
]

#: How the numbers are written. All of these appear in real client files, often
#: three of them inside one column.
def _money_cell(value: float, style: int, rng) -> object:
    if style == 0:
        return value
    if style == 1:
        return f"{value:,.2f}"
    if style == 2:
        return f"₹ {value:,.2f}"
    if style == 3:
        return f"{value:,.0f}/-"
    if style == 4 and value > 5000:
        return f"({value:,.0f})"          # a credit note, written as parentheses
    return value


def _date_cell(d: date, style: int) -> object:
    if style == 0:
        return d
    if style == 1:
        return d.strftime("%d-%m-%Y")
    if style == 2:
        return d.strftime("%d/%m/%y")
    if style == 3:
        return d.strftime("%d %b %Y")
    return d.strftime("%Y-%m-%d")


def bulk(count: int = 100, out: Path | None = None) -> list[tuple[str, str]]:
    """Generate ``count`` files spread across every shape a client sends.

    Roughly half Excel and half CSV, plus text exports, WhatsApp threads and a
    few deliberate wrecks — because a hundred clean files prove nothing that the
    first one did not.
    """
    out = out or (HERE / "samples" / "bulk")
    out.mkdir(parents=True, exist_ok=True)
    for old in out.glob("*"):
        if old.is_file():
            old.unlink()

    rng = random.Random(SEED + count)
    made: list[tuple[str, str]] = []

    # A different business per file, so party names are not all identical.
    firms = ["Ramu Stores", "Basavaraj Agri", "Krishna Traders", "Shetty Farms",
             "Patil Nursery", "Hanumanth & Sons", "Gouda Agencies", "Nandi Agro",
             "Bhat Hardware", "Malaprabha Traders", "Kittur Seeds", "Sankeshwar Feeds"]

    for n in range(count):
        kind = ("xlsx" if n % 10 < 4 else
                "csv" if n % 10 < 8 else
                "txt" if n % 10 == 8 else
                "wreck")
        month_back = n % 14
        when = ago(month_back * 30 + rng.randint(0, 27))
        stamp = when.strftime("%Y-%m")

        if kind == "wreck":
            # One in ten is broken on purpose. A corpus of only readable files
            # teaches nobody what a bad day looks like.
            name = f"{n:03d}-{stamp}-corrupt.xlsx"
            (out / name).write_bytes(rng.choice([
                b"%PDF-1.5\n% renamed by accident\n" + b"\x00" * 120,
                b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 120,  # old .xls
                b"<html><body>Export failed</body></html>",
            ]))
            made.append((name, "must fail cleanly and say what it really is"))
            continue

        rows = rng.randint(18, 60)
        headers = rng.choice(_HEADER_SHAPES)
        dstyle, mstyle = rng.randint(0, 4), rng.randint(0, 4)

        if kind == "csv":
            name = f"{n:03d}-{stamp}-sales.csv"
            with (out / name).open("w", newline="", encoding="utf-8") as fh:
                w = csv.writer(fh)
                w.writerow(headers)
                for i in range(rows):
                    sku, item, _c, _u, rate, _co = rng.choice(ITEMS)
                    qty = rng.choice([5, 10, 20, 25, 40, 50])
                    d = ago(month_back * 30 + rng.randint(0, 27))
                    line = _row(headers, d, dstyle, mstyle, rng.choice(firms),
                                sku, item, qty, rate, i)
                    w.writerow(line)
            made.append((name, f"CSV, {rows} rows, header shape {_HEADER_SHAPES.index(headers)}"))

        elif kind == "txt":
            name = f"{n:03d}-{stamp}-export.txt"
            with (out / name).open("w", encoding="utf-8") as fh:
                fh.write("\t".join(headers) + "\n")
                for i in range(rows):
                    sku, item, _c, _u, rate, _co = rng.choice(ITEMS)
                    qty = rng.choice([10, 20, 30])
                    d = ago(month_back * 30 + rng.randint(0, 27))
                    line = _row(headers, d, dstyle, mstyle, rng.choice(firms),
                                sku, item, qty, rate, i)
                    fh.write("\t".join(str(x) for x in line) + "\n")
            made.append((name, "tab-separated accounting export"))

        else:
            name = f"{n:03d}-{stamp}-book.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = rng.choice(["Sales", "Sheet1", "Sales Register", "Data"])
            # Junk above the header on roughly half of them.
            if rng.random() < 0.5:
                ws["A1"] = rng.choice(firms).upper() + " — MONTHLY STATEMENT"
                ws.merge_cells("A1:F1")
                ws["A1"].font = Font(bold=True, size=13)
                ws.append([])
            ws.append(headers)
            total = 0.0
            for i in range(rows):
                sku, item, _c, _u, rate, _co = rng.choice(ITEMS)
                qty = rng.choice([5, 10, 20, 25, 40])
                total += qty * rate
                d = ago(month_back * 30 + rng.randint(0, 27))
                ws.append(_row(headers, d, dstyle, mstyle, rng.choice(firms),
                               sku, item, qty, rate, i))
                if rng.random() < 0.04:
                    ws.append([])                     # a blank spacer mid-table
            if rng.random() < 0.45:
                ws.append([])
                ws.append(["", "", "Grand Total", "", "", total])

            # A stock sheet on some of them, so the corpus is not all sales.
            if rng.random() < 0.35:
                st = wb.create_sheet("Stock")
                st.append(rng.choice(_STOCK_SHAPES))
                for sku, item, _c, _u, rate, cost in ITEMS:
                    st.append([sku, item, rng.choice([0, 6, 14, 30, 70]),
                               rng.choice([10, 15, 20]), rate,
                               rng.choice(["Belagavi", "Hubballi"])][:st.max_column])
            for sheet in wb.worksheets:
                _autosize(sheet)
            wb.save(out / name)
            made.append((name, f"Excel, {rows} rows"))

    # A handful of WhatsApp threads, because that is a real intake route.
    for n in range(3):
        name = f"{count + n:03d}-whatsapp-thread.txt"
        lines = [f"{ago(n * 9 + 1).strftime('%d/%m/%Y')}, 9:0{n} am - "
                 f"Messages are end-to-end encrypted."]
        for _ in range(rng.randint(8, 16)):
            firm = rng.choice(firms)
            sku, item, _c, unit, rate, _co = rng.choice(ITEMS)
            qty = rng.choice([5, 10, 20, 30, 50])
            d = ago(n * 9 + rng.randint(0, 8))
            lines.append(f"{d.strftime('%d/%m/%Y')}, "
                         f"{rng.randint(8,18)}:{rng.randint(10,59)} - {firm}: "
                         + rng.choice([
                             f"{qty} {unit} {item.split()[0].lower()} beku",
                             f"send {qty} {item.split()[0].lower()}",
                             f"{item.split()[0].lower()} {qty} bags urgent",
                             f"rate enu {item.split()[0].lower()}?",
                             "payment done, check madi",
                         ]))
        (out / name).write_text("\n".join(lines), encoding="utf-8")
        made.append((name, "WhatsApp thread with orders in it"))

    (out / "README.md").write_text(
        "# Bulk corpus\n\n"
        f"{len(made)} files, generated by `python demo/make_samples.py --bulk "
        f"{count}`.\n\n"
        "Not nine chosen messes — a client's fourteen months of exports, written "
        "by whoever was on shift. Roughly 40% Excel, 40% CSV, 10% tab-separated, "
        "10% deliberately broken, plus WhatsApp threads.\n\n"
        "Header shape, date format and number format vary per file and sometimes "
        "within one. Some have junk rows above the header, some a Grand Total, "
        "some a blank spacer mid-table, some a stock sheet beside the sales.\n\n"
        "Use this to answer \"what happens when I send you everything I have\".\n",
        encoding="utf-8")
    return made


def _row(headers, d, dstyle, mstyle, party, sku, item, qty, rate, i):
    """One row shaped to whatever headers this file happens to use."""
    amount = qty * rate
    by_name = {
        "date": _date_cell(d, dstyle), "invoice date": _date_cell(d, dstyle),
        "bill dt": _date_cell(d, dstyle), "dt": _date_cell(d, dstyle),
        "sl": i + 1,
        "party": party, "customer name": party, "buyer": party,
        "party name": party, "customer": party,
        "item": item, "product": item, "item description": item,
        "particulars": item, "item name": item,
        "item code": sku,
        "qty": qty, "nos": qty, "quantity": qty,
        "rate": rate, "unit price": rate, "price": rate, "mrp": rate,
        "amount": _money_cell(amount, mstyle, None),
        "net amount": _money_cell(amount, mstyle, None),
        "value": _money_cell(amount, mstyle, None),
        "total": _money_cell(amount, mstyle, None),
        "remarks": "",
    }
    return [by_name.get(h.lower(), "") for h in headers]

if __name__ == "__main__":
    sys.exit(main())
