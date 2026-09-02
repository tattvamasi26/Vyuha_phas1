"""Manual books — for a business that has no spreadsheet at all.

A nursery owner writing sales in a notebook has the same questions as a
distributor with twelve monthly sheets: what is left, what sold, who owes me,
what is not moving. He just has nowhere to put the data.

So rather than build a second analytics engine for typed-in data, this module
keeps a small ledger and **writes it out as a workbook in exactly the shape the
engine already understands** — Sales Register, Stock Statement, Outstanding.
``pipeline.run()`` then treats it identically to an uploaded file. Typing a sale
and uploading a spreadsheet converge on the same dashboard, the same dead-stock
join, the same alerts.

The catalogue is deliberately generic: an ``Item`` is a plant, a bag of manure,
a pot or a packet of seeds. Stock decrements when a sale is recorded, so "what
is left" is answered by the ledger itself rather than by a stock-take.
"""

from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path

from .store import DATA

from . import atomic

BOOKS = DATA / "books"

#: Deliberately broad — a nursery, a manure dealer and a hardware shop all fit.
CATEGORIES = ["Plants", "Manure & Fertiliser", "Seeds", "Pots & Planters",
              "Tools", "Pesticides", "Soil & Compost", "Other"]
UNITS = ["piece", "kg", "bag", "packet", "litre", "tray", "dozen"]


def _today() -> str:
    return date.today().isoformat()


def _num(value, default: float = 0.0) -> float:
    try:
        return float(str(value).replace(",", "").replace("₹", "").strip())
    except (TypeError, ValueError):
        return default


@dataclass
class Item:
    """One thing he sells."""

    sku: str
    name: str
    category: str = "Other"
    unit: str = "piece"
    rate: float = 0.0            # what he sells it for
    cost: float = 0.0            # what it cost him — optional, drives margin
    stock_qty: float = 0.0       # what is left, right now
    reorder_level: float = 0.0
    added_at: str = field(default_factory=_today)
    #: Where this stock sits. Same contract as Sale.branch.
    branch: str = ""
    #: HSN/SAC code and GST rate, needed on a tax invoice line. Defaulted, so a
    #: business that does not charge GST is unaffected and its invoices simply
    #: show no tax columns. `gst_rate` is a percentage: 5 means 5%.
    hsn: str = ""
    gst_rate: float = 0.0

    @property
    def value(self) -> float:
        return self.stock_qty * (self.cost or self.rate)

    @property
    def low(self) -> bool:
        return self.reorder_level > 0 and self.stock_qty <= self.reorder_level


@dataclass
class Sale:
    """One line on one bill."""

    id: str
    date: str
    party: str
    sku: str
    item: str
    qty: float
    rate: float
    amount: float
    paid: bool = True            # False = sold on credit, becomes a receivable
    due_date: str = ""
    note: str = ""
    #: The buyer's WhatsApp number, digits with country code, no plus. Captured
    #: at the moment of sale because that is the only moment it is ever to hand
    #: — asking for it later means chasing a customer who has already left. It
    #: is what makes a receipt, and a payment reminder on a credit sale,
    #: possible at all.
    party_phone: str = ""
    receipt_sent: str = ""       # timestamp, so a receipt is never sent twice
    #: Which branch rang this up (``people.Branch.id``). Empty for a
    #: single-branch business and for every row written before branches
    #: existed, which is why people.performance() reports those under
    #: "Unassigned" rather than guessing.
    branch: str = ""
    #: Who made the sale (``people.Staff.id``). Without it "who is actually
    #: selling" is unanswerable, which is the first question an owner with
    #: staff asks and the one a branch total cannot answer.
    staff: str = ""


@dataclass
class Book:
    slug: str
    items: list[Item] = field(default_factory=list)
    sales: list[Sale] = field(default_factory=list)
    next_bill: int = 1

    # ---- derived views, answered without touching the engine ---------------
    @property
    def earned(self) -> float:
        return sum(s.amount for s in self.sales)

    @property
    def owed(self) -> float:
        return sum(s.amount for s in self.sales if not s.paid)

    @property
    def collected(self) -> float:
        return self.earned - self.owed

    @property
    def stock_value(self) -> float:
        return sum(i.value for i in self.items)

    @property
    def low_stock(self) -> list[Item]:
        return [i for i in self.items if i.low]

    @property
    def out_of_stock(self) -> list[Item]:
        return [i for i in self.items if i.stock_qty <= 0]

    @property
    def margin(self) -> float:
        """Profit on what has actually been sold, where a cost is known."""
        by_sku = {i.sku: i for i in self.items}
        total = 0.0
        for s in self.sales:
            item = by_sku.get(s.sku)
            if item and item.cost:
                total += (s.rate - item.cost) * s.qty
        return total

    def item(self, sku: str) -> Item | None:
        return next((i for i in self.items if i.sku == sku), None)

    def customers(self) -> list[str]:
        seen: dict[str, str] = {}
        for s in self.sales:
            seen.setdefault(s.party.strip().lower(), s.party.strip())
        return sorted(seen.values())

    def customer_phones(self) -> dict[str, str]:
        """Name to the last number we were given for them.

        Ramu buys every fortnight; asking for his number every fortnight is the
        kind of friction that stops a sale being recorded at all. Newest sale
        wins, so a corrected number replaces an old one.
        """
        found: dict[str, str] = {}
        for s in self.sales:                       # oldest first, later overwrite
            if s.party_phone:
                found[s.party.strip()] = s.party_phone
        return found

    def sale(self, sale_id: str) -> "Sale | None":
        return next((s for s in self.sales if s.id == sale_id), None)


# ------------------------------------------------------------------ persistence

def _path(slug: str) -> Path:
    BOOKS.mkdir(parents=True, exist_ok=True)
    return BOOKS / f"{slug}.json"


def load(slug: str) -> Book:
    path = _path(slug)
    if not path.exists():
        return Book(slug=slug)
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return Book(slug=slug)
    def build(cls, rows):
        # Drop unknown keys rather than raising: a data file written by a newer
        # build must not stop an older one from starting.
        return [cls(**{k: v for k, v in r.items() if k in cls.__dataclass_fields__})
                for r in rows]

    return Book(
        slug=slug,
        items=build(Item, raw.get("items", [])),
        sales=build(Sale, raw.get("sales", [])),
        next_bill=raw.get("next_bill", 1),
    )


def save(book: Book) -> None:
    atomic.write_json(_path(book.slug), {
        "items": [asdict(i) for i in book.items],
        "sales": [asdict(s) for s in book.sales],
        "next_bill": book.next_bill,
    })


# --------------------------------------------------------------------- editing

def make_sku(name: str, existing: list[Item]) -> str:
    base = re.sub(r"[^A-Z0-9]+", "-", name.upper()).strip("-")[:14] or "ITEM"
    taken = {i.sku for i in existing}
    if base not in taken:
        return base
    n = 2
    while f"{base}-{n}" in taken:
        n += 1
    return f"{base}-{n}"


def add_item(slug: str, name: str, category: str, unit: str, rate, cost,
             stock_qty, reorder_level) -> tuple[Book, str]:
    book = load(slug)
    name = name.strip()
    if not name:
        return book, "Give the item a name."

    existing = next((i for i in book.items if i.name.lower() == name.lower()), None)
    if existing:
        # Re-adding a known item is a restock, not a duplicate.
        existing.stock_qty += _num(stock_qty)
        if _num(rate):
            existing.rate = _num(rate)
        save(book)
        return book, f"Added {_num(stock_qty):g} to {existing.name}. Now {existing.stock_qty:g} in stock."

    book.items.append(Item(
        sku=make_sku(name, book.items), name=name,
        category=category or "Other", unit=unit or "piece",
        rate=_num(rate), cost=_num(cost),
        stock_qty=_num(stock_qty), reorder_level=_num(reorder_level),
    ))
    save(book)
    return book, f"{name} added."


def mark_receipt_sent(slug: str, sale_id: str) -> None:
    book = load(slug)
    sale = book.sale(sale_id)
    if sale is not None:
        sale.receipt_sent = datetime.now().isoformat(timespec="seconds")
        save(book)


def record_sale(slug: str, sku: str, party: str, qty, rate, when: str = "",
                paid: bool = True, due_date: str = "", note: str = "",
                party_phone: str = "", branch: str = "",
                staff: str = "") -> tuple[Book, str, bool]:
    book = load(slug)
    item = book.item(sku)
    if item is None:
        return book, "Pick an item from the list.", False

    qty_n, rate_n = _num(qty), _num(rate) or item.rate
    if qty_n <= 0:
        return book, "Quantity has to be more than zero.", False
    party = party.strip() or "Cash sale"

    warn = ""
    if qty_n > item.stock_qty:
        warn = (f" Note: only {item.stock_qty:g} {item.unit} were in stock, "
                f"so this now shows as {item.stock_qty - qty_n:g}.")

    bill = f"B-{book.next_bill:04d}"
    book.next_bill += 1
    book.sales.append(Sale(
        id=bill, date=when or _today(), party=party, sku=item.sku, item=item.name,
        qty=qty_n, rate=rate_n, amount=round(qty_n * rate_n, 2),
        paid=paid, due_date=due_date, note=note.strip(),
        party_phone=party_phone, branch=branch, staff=staff,
    ))
    item.stock_qty -= qty_n
    save(book)

    left = f"{item.stock_qty:g} {item.unit} left"
    return book, f"{bill}: {qty_n:g} × {item.name} to {party}. {left}.{warn}", True


def receive_stock(slug: str, sku: str, qty, cost="", note: str = "",
                  when: str = "") -> tuple[Book, str, bool]:
    """Stock coming in — a delivery, a return, a correction upward.

    Sales already move stock *out*; without this the only way a number ever
    went up was editing the item, which is why an inventory that starts
    accurate drifts within a week. A cost supplied here **replaces** the item's
    cost, because the price of the last delivery is the one that should drive
    margin from now on.
    """
    book = load(slug)
    item = book.item(sku)
    if item is None:
        return book, "Pick an item from the list.", False

    qty_n = _num(qty)
    if qty_n <= 0:
        return book, "How many came in? It has to be more than zero.", False

    before = item.stock_qty
    item.stock_qty += qty_n
    cost_n = _num(cost)
    priced = ""
    if cost_n > 0 and cost_n != item.cost:
        item.cost = cost_n
        priced = f" Cost now ₹{cost_n:,.0f}."
    save(book)
    return (book,
            f"{qty_n:g} {item.unit} of {item.name} received. "
            f"{before:g} → {item.stock_qty:g}.{priced}", True)


def adjust_stock(slug: str, sku: str, counted, note: str = "") -> tuple[Book, str, bool]:
    """Set stock to what was physically counted.

    Separate from receive_stock on purpose: "twelve arrived" and "there are
    twelve on the shelf" are different claims, and conflating them is how a
    stock-take quietly doubles the shelf.
    """
    book = load(slug)
    item = book.item(sku)
    if item is None:
        return book, "Pick an item from the list.", False

    counted_n = _num(counted, default=-1)
    if counted_n < 0:
        return book, "Enter the counted quantity.", False

    before = item.stock_qty
    item.stock_qty = counted_n
    save(book)
    delta = counted_n - before
    direction = "up" if delta > 0 else "down" if delta < 0 else "unchanged"
    return (book, f"{item.name} counted at {counted_n:g} {item.unit} "
                  f"({direction} {abs(delta):g} from {before:g}).", True)


def set_reorder(slug: str, levels: dict[str, str]) -> tuple[Book, str]:
    """Bulk-edit reorder levels — the whole table in one save.

    Setting them one at a time means nobody sets them at all, and a reorder
    level of zero is why ``Item.low`` never fires for most clients.
    """
    book = load(slug)
    changed = 0
    for sku, raw in levels.items():
        item = book.item(sku)
        if item is None:
            continue
        value = _num(raw, default=-1)
        if value < 0 or value == item.reorder_level:
            continue
        item.reorder_level = value
        changed += 1
    if changed:
        save(book)
        return book, f"Reorder level updated on {changed} item(s)."
    return book, "Nothing changed."


def set_branch(slug: str, sku: str, branch: str) -> tuple[Book, str]:
    """Move an item to a branch."""
    book = load(slug)
    item = book.item(sku)
    if item is None:
        return book, "That item is gone."
    item.branch = branch
    save(book)
    return book, f"{item.name} is now held at that branch."


def movement(book: Book, sku: str, limit: int = 20) -> list[dict]:
    """Everything that happened to one SKU, newest first.

    Only sales are dated individually, so receipts are not in the history yet —
    an honest gap, and the reason the console labels this "sales history"
    rather than "stock ledger".
    """
    rows = [{"date": s.date, "kind": "sold", "qty": -s.qty, "party": s.party,
             "amount": s.amount, "ref": s.id}
            for s in book.sales if s.sku == sku]
    rows.sort(key=lambda r: r["date"], reverse=True)
    return rows[:limit]


def delete_sale(slug: str, sale_id: str) -> tuple[Book, str]:
    book = load(slug)
    sale = next((s for s in book.sales if s.id == sale_id), None)
    if sale is None:
        return book, "That entry is already gone."
    book.sales = [s for s in book.sales if s.id != sale_id]
    item = book.item(sale.sku)
    if item:
        item.stock_qty += sale.qty          # put the stock back
    save(book)
    return book, f"{sale_id} removed and {sale.qty:g} put back into stock."


def delete_item(slug: str, sku: str) -> tuple[Book, str]:
    book = load(slug)
    item = book.item(sku)
    if item is None:
        return book, "That item is already gone."
    if any(s.sku == sku for s in book.sales):
        return book, f"{item.name} has sales recorded against it — those would lose their item."
    book.items = [i for i in book.items if i.sku != sku]
    save(book)
    return book, f"{item.name} removed."


def mark_paid(slug: str, sale_id: str) -> tuple[Book, str]:
    book = load(slug)
    sale = next((s for s in book.sales if s.id == sale_id), None)
    if sale is None:
        return book, "That entry is gone."
    sale.paid = True
    save(book)
    return book, f"{sale_id} marked paid."


# ----------------------------------------------------- hand it to the engine

def to_workbook(book: Book, out: Path) -> Path:
    """Write the ledger as a workbook the engine reads with no special case.

    The sheet names and headers deliberately match what ``schema.py`` already
    recognises, so a typed-in book and an uploaded file take identical paths
    through detect -> clean -> analyze -> report.
    """
    from openpyxl import Workbook

    out = Path(out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    sales = wb.active
    sales.title = "Sales Register"
    sales.append(["Date", "Invoice No.", "Party Name", "SKU", "Item", "Qty", "Rate", "Amount"])
    for s in book.sales:
        sales.append([s.date, s.id, s.party, s.sku, s.item, s.qty, s.rate, s.amount])

    stock = wb.create_sheet("Stock Statement")
    stock.append(["SKU", "Item", "Category", "Closing Stock", "Reorder Level", "Rate"])
    for i in book.items:
        stock.append([i.sku, i.name, i.category, i.stock_qty, i.reorder_level,
                      i.cost or i.rate])

    credit = [s for s in book.sales if not s.paid]
    if credit:
        due = wb.create_sheet("Outstanding")
        due.append(["Date", "Invoice No.", "Party Name", "Due Date", "Outstanding"])
        for s in credit:
            due.append([s.date, s.id, s.party, s.due_date or s.date, s.amount])

    for sheet in wb.worksheets:
        for col, cell in enumerate(sheet[1], start=1):
            width = max(12, min(28, len(str(cell.value)) + 6))
            sheet.column_dimensions[chr(64 + col)].width = width

    wb.save(str(out))
    return out


def summary(book: Book) -> dict:
    """The plain-language answers, for someone who will never open a dashboard."""
    sold_qty: dict[str, float] = {}
    for s in book.sales:
        sold_qty[s.sku] = sold_qty.get(s.sku, 0) + s.qty
    best = sorted(sold_qty.items(), key=lambda kv: kv[1], reverse=True)[:5]
    by_sku = {i.sku: i for i in book.items}

    return {
        "earned": book.earned,
        "collected": book.collected,
        "owed": book.owed,
        "margin": book.margin,
        "bills": len(book.sales),
        "customers": len(book.customers()),
        "items": len(book.items),
        "stock_value": book.stock_value,
        "low_stock": book.low_stock,
        "out_of_stock": book.out_of_stock,
        "best_sellers": [(by_sku[k].name if k in by_sku else k, v) for k, v in best],
        "never_sold": [i for i in book.items if i.sku not in sold_qty and i.stock_qty > 0],
        "last_updated": max((s.date for s in book.sales), default=""),
    }


def bill_text(book: Book, sale_id: str, business: str) -> str:
    """A plain-text bill he can send on WhatsApp."""
    sale = next((s for s in book.sales if s.id == sale_id), None)
    if sale is None:
        return ""
    when = datetime.fromisoformat(sale.date).strftime("%d %b %Y") if sale.date else ""
    status = "PAID" if sale.paid else "DUE"
    return (f"*{business}*\n"
            f"Bill {sale.id} · {when}\n\n"
            f"{sale.party}\n\n"
            f"{sale.item}\n{sale.qty:g} × ₹{sale.rate:,.0f} = *₹{sale.amount:,.0f}*\n\n"
            f"Status: {status}")
